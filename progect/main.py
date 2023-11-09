import openpyxl
import os
# Объявляем константу в переменньй: DIR будет находиться, путь дирректории где запущен будет скрипт
DIR = os.path.abspath(os.curdir)

# цикл для поиска документов в формате xlxs и переименования в нужный формат - имя
for root, dirs, files in os.walk(DIR):
    for file in files:
        if file.endswith('.xlsx'):
            os.rename(file, 'ошибки сотрудников.xlsx')

wb = openpyxl.load_workbook('ошибки сотрудников.xlsx')
sheets = wb.active
# цикл для проставления дальнейших действий
for i in range(2, sheets.max_row+1):
    b = sheets[i][9].value
    if b =='Неверный АРМ':
        sheets.cell(row=i, column=11).value = 'зайти под верным АРМ и переподписать'
    elif b == 'СНИЛС':
        sheets.cell(row=i, column=11).value = 'Запросить СНИЛС пациента'
    elif b == 'На дату создания документа для указанного вида требуется как минимум 1 подпись роли [DOCTOR]':
        sheets.cell(row=i, column=11).value = 'При подписании выбрать правильную роль: Врач'
    else:
        sheets.cell(row=i, column=11).value = 'Тут ничего не нужно'
# сохранение файла
wb.save('ошибки сотрудников.xlsx')
